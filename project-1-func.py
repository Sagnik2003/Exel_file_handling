import openpyxl as xl
from openpyxl.chart import BarChart, Reference

filename = input('filename : ')


def process_workbook(filename):
    wb = xl.load_workbook(filename)
    sheet = wb['Sheet1']
    cell = sheet['a1']
    cell = sheet.cell(1, 1)  # both are same
    print(sheet.max_row)
    # max_row means ,how many rows are there in the exel sheet
    # min_column means, how many columns are there in theexel sheeet
    print(sheet.max_column)

    cell = sheet.cell(1, 4)
    cell.value = 'price2'

    for row in range(2, sheet.max_row + 1):
        cell = sheet.cell(row, 3)
        corrected_price = cell.value*0.9
        corrected_price_cell = sheet.cell(row, 4)
        corrected_price_cell.value = corrected_price

    values = Reference(sheet,
                       min_row=2,
                       max_row=sheet.max_row,
                       min_col=4,
                       max_col=4)

    chart = BarChart()
    chart.add_data(values)
    sheet.add_chart(chart, 'e2')

    wb.save(filename)
